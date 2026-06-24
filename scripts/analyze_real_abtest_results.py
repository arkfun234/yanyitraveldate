#!/usr/bin/env python3
"""Analyze real A/B test participant JSON exports.

This A/B Test analysis is kept as a preliminary evaluation workflow.
The current research direction is A+B integrated recommendation, where
A provides the traveler-type framework and B provides free-text clustering
signals for explanation, reproducibility, and diversity.

The script copies participant exports from Downloads into the project,
analyzes the project copies, and creates Excel and Markdown reports.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

try:
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    print(
        "Required packages are missing. Install them with:\n"
        "py -m pip install pandas openpyxl",
        file=sys.stderr,
    )
    raise SystemExit(1) from exc


DEFAULT_SOURCE_DIR = Path(r"C:\Users\QQQ\Downloads")
DEFAULT_PATTERN = "ab_recommendation_with_ai_plan_*.json"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PARTICIPANT_DIR = PROJECT_ROOT / "abtest_results" / "real_participants"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data" / "real_abtest_analysis"

PLAN1_METHOD = "A_handcrafted_matrix"
PLAN2_METHOD = "B_auto_clustering_baseline"

SCORE_KEYS = [
    "answer_fit",
    "visit_interest",
    "reason_clarity",
    "personalized_feeling",
    "overall_satisfaction",
]
SCORE_LABELS = {
    "answer_fit": "自分の回答に合っている",
    "visit_interest": "行ってみたいと思う",
    "reason_clarity": "推薦理由が分かりやすい",
    "personalized_feeling": "個人化されていると感じる",
    "overall_satisfaction": "全体的に満足できる",
    "five_item_average": "5項目平均",
}
CHOICE_TO_METHOD = {
    "推薦案1": PLAN1_METHOD,
    "推薦案2": PLAN2_METHOD,
    "どちらとも言えない": "neutral",
}
CHOICE_DISPLAY = {
    PLAN1_METHOD: "A案",
    PLAN2_METHOD: "B案",
    "neutral": "どちらとも言えない",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze real A/B test participant JSON files."
    )
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=DEFAULT_SOURCE_DIR,
        help=f"Directory containing source JSON files (default: {DEFAULT_SOURCE_DIR})",
    )
    parser.add_argument(
        "--pattern",
        default=DEFAULT_PATTERN,
        help=f"Source filename pattern (default: {DEFAULT_PATTERN})",
    )
    parser.add_argument(
        "--participant-dir",
        type=Path,
        default=DEFAULT_PARTICIPANT_DIR,
        help="Project directory where source JSON files are copied.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for Excel and Markdown reports.",
    )
    return parser.parse_args()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def copy_source_files(
    source_dir: Path, pattern: str, participant_dir: Path
) -> tuple[list[Path], int, int]:
    participant_dir.mkdir(parents=True, exist_ok=True)
    source_files = sorted(source_dir.glob(pattern))
    copied_count = 0
    unchanged_count = 0

    for source in source_files:
        if not source.is_file():
            continue
        destination = participant_dir / source.name
        if destination.exists() and sha256(source) == sha256(destination):
            unchanged_count += 1
            continue
        shutil.copy2(source, destination)
        copied_count += 1

    return source_files, copied_count, unchanged_count


def first_not_none(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None


def canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )


def display_json(value: Any) -> str:
    if value is None:
        return ""
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def spot_names(spots: Any) -> str:
    if not isinstance(spots, list):
        return ""
    names = []
    for spot in spots:
        if isinstance(spot, dict):
            name = spot.get("name")
            if name:
                names.append(str(name))
        elif spot:
            names.append(str(spot))
    return " / ".join(names)


def numeric_score(scores: dict[str, Any], key: str) -> float | None:
    value = scores.get(key)
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def average_scores(scores: dict[str, Any]) -> float | None:
    values = [numeric_score(scores, key) for key in SCORE_KEYS]
    valid_values = [value for value in values if value is not None]
    if not valid_values:
        return None
    return sum(valid_values) / len(valid_values)


def load_participant(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig") as handle:
        data = json.load(handle)

    payload = data.get("original_ab_evaluation_payload") or {}
    profile = data.get("user_profile") or {}
    cluster = profile.get("cluster") or {}
    ai_plan = data.get("local_ai_plan") or {}

    questionnaire = first_not_none(
        data.get("questionnaire_answers"),
        payload.get("questionnaire_answers"),
        {},
    )
    plan1_scores = first_not_none(data.get("plan1_scores"), payload.get("plan1_scores"), {})
    plan2_scores = first_not_none(data.get("plan2_scores"), payload.get("plan2_scores"), {})
    plan1_spots = first_not_none(
        data.get("plan1_recommendations"),
        payload.get("recommendation_plan_1_spots"),
        [],
    )
    plan2_spots = first_not_none(
        data.get("plan2_recommendations"),
        payload.get("recommendation_plan_2_spots"),
        [],
    )
    comparison_choice = first_not_none(
        data.get("comparison_choice"),
        payload.get("comparison_choice"),
        "",
    )
    free_comment = first_not_none(
        data.get("free_comment"),
        payload.get("free_comment"),
        payload.get("comment"),
        "",
    )
    timestamp = first_not_none(
        payload.get("timestamp"),
        data.get("timestamp"),
        data.get("saved_at"),
        "",
    )

    a_average = average_scores(plan1_scores)
    b_average = average_scores(plan2_scores)
    difference = (
        a_average - b_average
        if a_average is not None and b_average is not None
        else None
    )
    selected_method = CHOICE_TO_METHOD.get(
        str(comparison_choice), f"unmapped:{comparison_choice}"
    )

    row: dict[str, Any] = {
        "file_name": path.name,
        "timestamp": timestamp,
        "questionnaire_answers": display_json(questionnaire),
        "user_cluster_id": cluster.get("id"),
        "user_cluster_name": cluster.get("name", ""),
        "companion": profile.get("companion", ""),
        "season": profile.get("season", ""),
        "visited_before": profile.get("visited_before"),
        "visited_places": display_json(profile.get("visited_places"))
        if isinstance(profile.get("visited_places"), (list, dict))
        else first_not_none(profile.get("visited_places"), ""),
        "plan1_method": PLAN1_METHOD,
        "plan2_method": PLAN2_METHOD,
        "plan1_spots": spot_names(plan1_spots),
        "plan2_spots": spot_names(plan2_spots),
        "comparison_choice": comparison_choice,
        "free_comment": free_comment,
        "local_ai_plan.ok": ai_plan.get("ok"),
        "local_ai_plan.model": ai_plan.get("model", ""),
        "local_ai_plan.generated_at": ai_plan.get("generated_at", ""),
        "A_5item_average": a_average,
        "B_5item_average": b_average,
        "score_difference_A_minus_B": difference,
        "selected_method": selected_method,
    }

    for key in SCORE_KEYS:
        row[f"plan1_scores.{key}"] = numeric_score(plan1_scores, key)
        row[f"plan2_scores.{key}"] = numeric_score(plan2_scores, key)

    if isinstance(questionnaire, dict):
        for question in sorted(
            questionnaire,
            key=lambda item: (
                int(str(item)[1:]) if str(item).startswith("Q") and str(item)[1:].isdigit() else 9999,
                str(item),
            ),
        ):
            row[f"questionnaire.{question}"] = questionnaire[question]

    duplicate_key = canonical_json(
        {
            "questionnaire_answers": questionnaire,
            "plan1_scores": plan1_scores,
            "plan2_scores": plan2_scores,
            "comparison_choice": comparison_choice,
        }
    )
    return row, {"duplicate_key": duplicate_key}


def mark_duplicates(
    rows: list[dict[str, Any]], metadata: list[dict[str, Any]]
) -> None:
    keys = [item["duplicate_key"] for item in metadata]
    group_counts = Counter(keys)
    group_numbers: dict[str, int] = {}
    next_group = 1
    seen: Counter[str] = Counter()

    for row, key in zip(rows, keys):
        count = group_counts[key]
        if count > 1 and key not in group_numbers:
            group_numbers[key] = next_group
            next_group += 1
        seen[key] += 1
        row["likely_duplicate_group"] = (
            f"DUP-{group_numbers[key]:03d}" if count > 1 else ""
        )
        row["duplicate_group_size"] = count
        row["is_in_duplicate_group"] = count > 1
        row["exclude_as_later_duplicate"] = count > 1 and seen[key] > 1
        row["duplicate_status"] = (
            "重複候補（代表行）"
            if count > 1 and seen[key] == 1
            else "重複候補（除外対象）"
            if count > 1
            else "重複なし"
        )


def score_summary(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    datasets = [
        ("全ファイル", df),
        ("重複除外", df.loc[~df["exclude_as_later_duplicate"]].copy()),
    ]
    for dataset_name, subset in datasets:
        for key in SCORE_KEYS:
            a_value = subset[f"plan1_scores.{key}"].mean()
            b_value = subset[f"plan2_scores.{key}"].mean()
            records.append(
                {
                    "集計対象": dataset_name,
                    "評価項目": SCORE_LABELS[key],
                    "A方案平均分": a_value,
                    "B方案平均分": b_value,
                    "差分 A-B": a_value - b_value,
                    "回答数": len(subset),
                }
            )
        a_average = subset["A_5item_average"].mean()
        b_average = subset["B_5item_average"].mean()
        records.append(
            {
                "集計対象": dataset_name,
                "評価項目": SCORE_LABELS["five_item_average"],
                "A方案平均分": a_average,
                "B方案平均分": b_average,
                "差分 A-B": a_average - b_average,
                "回答数": len(subset),
            }
        )
    return pd.DataFrame(records)


def choice_summary(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    datasets = [
        ("全ファイル", df),
        ("重複除外", df.loc[~df["exclude_as_later_duplicate"]].copy()),
    ]
    choices = [
        (PLAN1_METHOD, "A案 selected count"),
        (PLAN2_METHOD, "B案 selected count"),
        ("neutral", "どちらとも言えない count"),
    ]
    for dataset_name, subset in datasets:
        total = len(subset)
        for method, label in choices:
            count = int((subset["selected_method"] == method).sum())
            records.append(
                {
                    "集計対象": dataset_name,
                    "選択": label,
                    "selected_method": method,
                    "count": count,
                    "percentage": count / total if total else 0,
                }
            )
        unmapped = int(subset["selected_method"].astype(str).str.startswith("unmapped:").sum())
        if unmapped:
            records.append(
                {
                    "集計対象": dataset_name,
                    "選択": "未マッピング count",
                    "selected_method": "unmapped",
                    "count": unmapped,
                    "percentage": unmapped / total if total else 0,
                }
            )
    return pd.DataFrame(records)


def ai_plan_summary(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    datasets = [
        ("全ファイル", df),
        ("重複除外", df.loc[~df["exclude_as_later_duplicate"]].copy()),
    ]
    for dataset_name, subset in datasets:
        success_mask = subset["local_ai_plan.ok"].fillna(False).astype(bool)
        records.extend(
            [
                {"集計対象": dataset_name, "指標": "total files", "値": len(subset)},
                {
                    "集計対象": dataset_name,
                    "指標": "AI plan generation success count",
                    "値": int(success_mask.sum()),
                },
                {
                    "集計対象": dataset_name,
                    "指標": "failure count",
                    "値": int((~success_mask).sum()),
                },
            ]
        )
        model_counts = subset["local_ai_plan.model"].replace("", "(model未記録)").fillna(
            "(model未記録)"
        ).value_counts()
        for model, count in model_counts.items():
            records.append(
                {
                    "集計対象": dataset_name,
                    "指標": f"model count: {model}",
                    "値": int(count),
                }
            )
    return pd.DataFrame(records)


def duplicate_check(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "file_name",
        "timestamp",
        "likely_duplicate_group",
        "duplicate_group_size",
        "is_in_duplicate_group",
        "exclude_as_later_duplicate",
        "duplicate_status",
        "questionnaire_answers",
        *[f"plan1_scores.{key}" for key in SCORE_KEYS],
        *[f"plan2_scores.{key}" for key in SCORE_KEYS],
        "comparison_choice",
    ]
    return df[columns].sort_values(
        ["is_in_duplicate_group", "likely_duplicate_group", "timestamp"],
        ascending=[False, True, True],
    )


def report_ready_table(
    df: pd.DataFrame, scores: pd.DataFrame, choices: pd.DataFrame
) -> pd.DataFrame:
    effective = df.loc[~df["exclude_as_later_duplicate"]]
    rows = [
        {
            "区分": "回答数",
            "指標": "有効回答数",
            "全ファイル": len(df),
            "重複除外": len(effective),
            "備考": f"後続の重複候補 {int(df['exclude_as_later_duplicate'].sum())} 件を除外",
        }
    ]
    for label in SCORE_LABELS.values():
        all_row = scores[(scores["集計対象"] == "全ファイル") & (scores["評価項目"] == label)].iloc[0]
        dedup_row = scores[
            (scores["集計対象"] == "重複除外") & (scores["評価項目"] == label)
        ].iloc[0]
        rows.extend(
            [
                {
                    "区分": "平均評価",
                    "指標": f"{label}：A案",
                    "全ファイル": all_row["A方案平均分"],
                    "重複除外": dedup_row["A方案平均分"],
                    "備考": "5点満点",
                },
                {
                    "区分": "平均評価",
                    "指標": f"{label}：B案",
                    "全ファイル": all_row["B方案平均分"],
                    "重複除外": dedup_row["B方案平均分"],
                    "備考": "5点満点",
                },
                {
                    "区分": "平均評価",
                    "指標": f"{label}：差分 A-B",
                    "全ファイル": all_row["差分 A-B"],
                    "重複除外": dedup_row["差分 A-B"],
                    "備考": "正の値はA案優位",
                },
            ]
        )
    for method in [PLAN1_METHOD, PLAN2_METHOD, "neutral"]:
        all_row = choices[
            (choices["集計対象"] == "全ファイル")
            & (choices["selected_method"] == method)
        ].iloc[0]
        dedup_row = choices[
            (choices["集計対象"] == "重複除外")
            & (choices["selected_method"] == method)
        ].iloc[0]
        rows.append(
            {
                "区分": "選択結果",
                "指標": CHOICE_DISPLAY[method],
                "全ファイル": f"{int(all_row['count'])}件 ({all_row['percentage']:.1%})",
                "重複除外": f"{int(dedup_row['count'])}件 ({dedup_row['percentage']:.1%})",
                "備考": "",
            }
        )
    return pd.DataFrame(rows)


def markdown_table(headers: list[str], rows: Iterable[Iterable[Any]]) -> str:
    def clean(value: Any) -> str:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        return str(value).replace("|", "\\|").replace("\n", " ")

    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    lines.extend("| " + " | ".join(clean(value) for value in row) + " |" for row in rows)
    return "\n".join(lines)


def interpretation_text(
    df: pd.DataFrame, scores: pd.DataFrame, choices: pd.DataFrame
) -> str:
    effective = df.loc[~df["exclude_as_later_duplicate"]]
    five = scores[
        (scores["集計対象"] == "重複除外")
        & (scores["評価項目"] == SCORE_LABELS["five_item_average"])
    ].iloc[0]
    a_avg = five["A方案平均分"]
    b_avg = five["B方案平均分"]
    diff = five["差分 A-B"]

    if abs(diff) < 0.10:
        score_statement = "5項目平均ではA案とB案の差は小さく、概ね同程度の評価でした。"
    elif diff > 0:
        score_statement = f"5項目平均ではA案がB案を{diff:.2f}点上回りました。"
    else:
        score_statement = f"5項目平均ではB案がA案を{abs(diff):.2f}点上回りました。"

    effective_choices = choices[choices["集計対象"] == "重複除外"].set_index(
        "selected_method"
    )
    a_count = int(effective_choices.loc[PLAN1_METHOD, "count"])
    b_count = int(effective_choices.loc[PLAN2_METHOD, "count"])
    neutral_count = int(effective_choices.loc["neutral", "count"])
    success_count = int(
        effective["local_ai_plan.ok"].fillna(False).astype(bool).sum()
    )

    return (
        f"重複候補の後続回答を除いた有効回答は{len(effective)}件です。"
        f"{score_statement}"
        f"選択結果はA案{a_count}件、B案{b_count}件、中立{neutral_count}件でした。"
        f"AI旅行計画の生成成功は{success_count}/{len(effective)}件です。"
        f"平均点はA案{a_avg:.2f}点、B案{b_avg:.2f}点であり、"
        "現時点では標本数が限られるため、方向性を示す記述的結果として扱うのが適切です。"
    )


def write_markdown(
    output_path: Path,
    df: pd.DataFrame,
    scores: pd.DataFrame,
    choices: pd.DataFrame,
) -> None:
    duplicate_rows = int(df["exclude_as_later_duplicate"].sum())
    effective_count = len(df) - duplicate_rows

    score_rows = []
    for dataset_name in ["全ファイル", "重複除外"]:
        subset = scores[scores["集計対象"] == dataset_name]
        for _, row in subset.iterrows():
            score_rows.append(
                [
                    dataset_name,
                    row["評価項目"],
                    f"{row['A方案平均分']:.2f}",
                    f"{row['B方案平均分']:.2f}",
                    f"{row['差分 A-B']:+.2f}",
                ]
            )

    choice_rows = []
    for _, row in choices.iterrows():
        choice_rows.append(
            [
                row["集計対象"],
                row["選択"],
                int(row["count"]),
                f"{row['percentage']:.1%}",
            ]
        )

    ai_rows = []
    for dataset_name, subset in [
        ("全ファイル", df),
        ("重複除外", df.loc[~df["exclude_as_later_duplicate"]]),
    ]:
        success = int(subset["local_ai_plan.ok"].fillna(False).astype(bool).sum())
        models = subset["local_ai_plan.model"].replace("", "(model未記録)").fillna(
            "(model未記録)"
        ).value_counts()
        ai_rows.append(
            [
                dataset_name,
                len(subset),
                success,
                len(subset) - success,
                ", ".join(f"{model}: {count}" for model, count in models.items()),
            ]
        )

    content = f"""# 実参加者 A/B テスト分析サマリー

## データ概要

- 読み込んだJSONファイル数: {len(df)}
- 重複の可能性が高い行数（先頭回答を除く）: {duplicate_rows}
- 重複除外後の有効回答数: {effective_count}

## A/B 平均評価

{markdown_table(
    ["集計対象", "評価項目", "A方案平均分", "B方案平均分", "差分 A-B"],
    score_rows,
)}

## 選択件数

{markdown_table(["集計対象", "選択", "件数", "割合"], choice_rows)}

## AI旅行計画の生成状況

{markdown_table(
    ["集計対象", "総数", "成功", "失敗", "モデル別件数"],
    ai_rows,
)}

## 会議報告用の短い解釈

{interpretation_text(df, scores, choices)}

## 重複判定方法

questionnaire_answers、plan1_scores、plan2_scores、comparison_choice がすべて一致する回答を同一の重複候補グループとしました。重複除外集計では、各グループの最初の回答を残し、それ以降の回答を除外しています。
"""
    output_path.write_text(content, encoding="utf-8")


def format_workbook(excel_path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    duplicate_fill = PatternFill("solid", fgColor="FCE4D6")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[:200]
            )
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)

        if worksheet.title == "duplicate_check":
            headers = {cell.value: cell.column for cell in worksheet[1]}
            status_column = headers.get("is_in_duplicate_group")
            if status_column:
                for row_number in range(2, worksheet.max_row + 1):
                    if worksheet.cell(row=row_number, column=status_column).value:
                        for cell in worksheet[row_number]:
                            cell.fill = duplicate_fill

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for sheet_name in ["score_summary", "choice_summary", "report_ready_table"]:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00%"
                    if sheet_name != "choice_summary" or worksheet.cell(1, cell.column).value != "percentage":
                        cell.number_format = "0.00"

    choice_sheet = workbook["choice_summary"]
    choice_headers = {cell.value: cell.column for cell in choice_sheet[1]}
    percentage_column = choice_headers.get("percentage")
    if percentage_column:
        for row_number in range(2, choice_sheet.max_row + 1):
            choice_sheet.cell(row=row_number, column=percentage_column).number_format = "0.0%"

    workbook.save(excel_path)


def main() -> int:
    args = parse_args()
    source_dir = args.source_dir.expanduser().resolve()
    participant_dir = args.participant_dir.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()

    if not source_dir.exists():
        print(f"Source directory does not exist: {source_dir}", file=sys.stderr)
        return 2

    source_files, copied_count, unchanged_count = copy_source_files(
        source_dir, args.pattern, participant_dir
    )
    if not source_files:
        print(
            f"No files matched {source_dir / args.pattern}. No reports were created.",
            file=sys.stderr,
        )
        return 3

    analysis_files = sorted(participant_dir.glob(args.pattern))
    rows: list[dict[str, Any]] = []
    metadata: list[dict[str, Any]] = []
    errors: list[str] = []
    for path in analysis_files:
        try:
            row, meta = load_participant(path)
            rows.append(row)
            metadata.append(meta)
        except (OSError, UnicodeError, json.JSONDecodeError, TypeError, ValueError) as exc:
            errors.append(f"{path.name}: {type(exc).__name__}: {exc}")

    if not rows:
        print("No valid participant JSON files could be analyzed.", file=sys.stderr)
        for error in errors:
            print(f"  {error}", file=sys.stderr)
        return 4

    mark_duplicates(rows, metadata)
    participants_df = pd.DataFrame(rows)

    preferred_columns = [
        "file_name",
        "timestamp",
        "questionnaire_answers",
        *[column for column in participants_df.columns if column.startswith("questionnaire.Q")],
        "user_cluster_id",
        "user_cluster_name",
        "companion",
        "season",
        "visited_before",
        "visited_places",
        "plan1_method",
        "plan2_method",
        "plan1_spots",
        "plan2_spots",
        *[f"plan1_scores.{key}" for key in SCORE_KEYS],
        *[f"plan2_scores.{key}" for key in SCORE_KEYS],
        "A_5item_average",
        "B_5item_average",
        "score_difference_A_minus_B",
        "comparison_choice",
        "selected_method",
        "free_comment",
        "local_ai_plan.ok",
        "local_ai_plan.model",
        "local_ai_plan.generated_at",
        "likely_duplicate_group",
        "duplicate_group_size",
        "is_in_duplicate_group",
        "exclude_as_later_duplicate",
        "duplicate_status",
    ]
    participants_df = participants_df[
        [column for column in preferred_columns if column in participants_df.columns]
    ]

    scores_df = score_summary(participants_df)
    choices_df = choice_summary(participants_df)
    ai_df = ai_plan_summary(participants_df)
    duplicates_df = duplicate_check(participants_df)
    report_df = report_ready_table(participants_df, scores_df, choices_df)

    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / "real_abtest_summary.xlsx"
    markdown_path = output_dir / "real_abtest_summary.md"
    csv_dir = output_dir / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)

    csv_tables = {
        "participants_raw.csv": participants_df,
        "score_summary.csv": scores_df,
        "choice_summary.csv": choices_df,
        "ai_plan_summary.csv": ai_df,
        "duplicate_check.csv": duplicates_df,
        "report_ready_table.csv": report_df,
    }
    for file_name, table in csv_tables.items():
        table.to_csv(csv_dir / file_name, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        participants_df.to_excel(writer, sheet_name="participants_raw", index=False)
        scores_df.to_excel(writer, sheet_name="score_summary", index=False)
        choices_df.to_excel(writer, sheet_name="choice_summary", index=False)
        ai_df.to_excel(writer, sheet_name="ai_plan_summary", index=False)
        duplicates_df.to_excel(writer, sheet_name="duplicate_check", index=False)
        report_df.to_excel(writer, sheet_name="report_ready_table", index=False)

    format_workbook(excel_path)
    write_markdown(markdown_path, participants_df, scores_df, choices_df)

    print(f"Source files matched: {len(source_files)}")
    print(f"JSON files copied or updated: {copied_count}")
    print(f"JSON files already identical: {unchanged_count}")
    print(f"JSON files analyzed: {len(participants_df)}")
    print(
        "Likely duplicate rows excluded: "
        f"{int(participants_df['exclude_as_later_duplicate'].sum())}"
    )
    print(f"Excel report: {excel_path}")
    print(f"Markdown report: {markdown_path}")
    print(f"CSV reports: {csv_dir}")
    if errors:
        print(f"Files skipped due to errors: {len(errors)}")
        for error in errors:
            print(f"  {error}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
